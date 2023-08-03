import PySimpleGUI as sg
from datetime import datetime, timedelta, date
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import string
import os
import re
import csv
import itertools
from difflib import SequenceMatcher
import math

settingsPath = 'settings.json'
uppercase = list(string.ascii_uppercase)
currentyear = datetime.today().year
defaultStart = f'{currentyear}-01-01'
defaultEnd = f'{currentyear}-01-31'

def stringSimilar(a, b):
    conf = 0
    for x, y in zip(a, b):
        if x != y:
            break
        conf += 1
    if conf >= 15:
        return 1
    conf = (conf - 8) ** 3 / 2000 + 0.03 * conf # Math witchcraft
    conf += SequenceMatcher(None, a, b).ratio() * 0.75
    return conf > 0.5

def defaultSettings():
    settings['In'] = './input.csv'
    settings['Out'] = './output.xlsx'
    settings['Columns'] = {'Post': 'B', 'Description': 'D', 'Debit': 'F', 'Credit': 'G'}
    settings['HasHeader'] = True
    settings['Periods'] = []
    settings['DateFormat'] = '%Y %m %d'
    settings['Exclusions'] = ''
    settings['PositiveDebits'] = False
    settings['SortDesc'] = False

def tryFloat(input):
    try:
        return float(input)
    except:
        return 0

def parse():
    # Checks if data are proper parameters
    columnIndices = []
    for c in settings['Columns'].values():
        columnIndices.append(ord(c) - ord('A'))
    if not os.path.isfile(settings['In']):
        return 'Cannot find the input file.'
    periods = []
    for period in settings['Periods']:
        try:
            start = datetime.strptime(period[0], '%Y-%m-%d')
        except:
            return f'{period[0]} could not be interpreted as a real date.'
        try:
            end = datetime.strptime(period[1], '%Y-%m-%d')
        except:
            return f'{period[1]} could not be interpreted as a real date.'
        if start > end:
            return f'Start date {start} is before end date {end}'
        periods.append((start, end))

    # Reads and parses the raw data
    bigList = [{'Debit': {}, 'Credit': {}} for i in periods]
    with open(settings['In']) as f:
        file = [line for line in csv.reader(f)]
    testLine = file[1 if settings['HasHeader'] else 0]
    for letter, index in zip(settings['Columns'].values(), columnIndices):
        if index < 0:
            return f'{letter} could not be interpreted as a real column.'
        if len(testLine) <= index:
            return f'Could not find the {letter} column.'

    excludeWords = [word.strip() for word in settings['Exclusions'].lower().split(',') if word and not word.isspace()]
    for line in file[1 if settings['HasHeader'] else 0:]:
        if len(line) > 1:
            try:
                date = datetime.strptime(re.sub('[-/:;]', ' ', line[columnIndices[0]]), settings['DateFormat'])
            except:
                return 'The posted dates do not match the date format.'
            desc = line[columnIndices[1]]
            debit = abs(tryFloat(line[columnIndices[2]]))
            credit = abs(tryFloat(line[columnIndices[3]]))

            # Categorize by period
            index = -1
            for i, period in enumerate(periods):
                if period[0] <= date <= period[1]:
                    index = i
                    break
            if index != -1:
                creditList = bigList[index]

                # Categorize by credit vs debit
                periodList = creditList['Debit' if debit else 'Credit']

                # Categorize by transaction client
                descLower = desc.lower()
                similar = ''
                for word in periodList.keys():
                    if stringSimilar(word, descLower):
                        similar = word
                        break

                if similar:
                    client = periodList[word]
                    if isinstance(client[0], list):
                        client.append([date, desc, debit, credit])
                    else:
                        client[2] += debit
                        client[3] += credit
                else:
                    newCells = [date, desc, debit, credit]
                    # Creates a list to store line items if it contains an excluded word
                    if any([word in descLower for word in excludeWords]):
                        newCells = [newCells]
                    periodList[descLower] = newCells


    # Formats data into an excel file
    wb = Workbook()
    ws = wb.active

    def formatRange(a, b, c, d, mode, format):
        for c in range(ord(a), ord(c) + 1):
            ch = chr(c)
            for i in range(b, d + 1):
                match mode:
                    case 'font':
                        ws[f'{ch}{i}'].font = format
                    case 'number':
                        ws[f'{ch}{i}'].number_format = format
                    case 'fill':
                        ws[f'{ch}{i}'].fill = format

    lineNumber = 1
    anchor = -1
    header = ['Last Posted Date', 'Description', 'Debit', 'Credit']
    for i, creditList in enumerate(bigList):
        # Add new rows
        startBound = lineNumber + 6
        endBound = lineNumber + 5 + len(creditList['Credit']) + len(creditList['Debit'])
        for row in itertools.chain(creditList['Credit'].values(), creditList['Debit'].values()):
            if isinstance(row[0], list):
                endBound += len(row) - 1
        ws.append([f'{periods[i][0].strftime("%b %d, %Y")} - {periods[i][1].strftime("%b %d, %Y")}\n'])
        ws.append(['Previous Balance', 0 if anchor == -1 else f'=B{anchor}'])
        ws.append(['Total Debits', f'={"" if settings["PositiveDebits"] else "-"}SUM(C{startBound}:C{endBound})'])
        ws.append(['Total Credits', f'={"-" if settings["PositiveDebits"] else ""}SUM(D{startBound}:D{endBound})'])
        ws.append(['New Balance', f'=SUM(B{lineNumber + 1}:B{lineNumber + 3})'])
        ws.append(header)
        if settings['SortDesc']:
            cList = [value for key,value in sorted(creditList['Credit'].items())]
            dList = [value for key,value in sorted(creditList['Debit'].items())]
        else:
            cList = creditList['Credit'].values()
            dList = creditList['Debit'].values()
        for row in itertools.chain(cList, dList):
            if isinstance(row[0], list):
                for r in row:
                    ws.append(r)
            else:
                ws.append(row)
        ws.append([])

        # Format
        formatRange('A', lineNumber, 'D', lineNumber + 5, 'font', Font(bold=True))
        formatRange('A', lineNumber, 'D', lineNumber + 4, 'fill', PatternFill("solid", fgColor="DDDDDD"))
        formatRange('B', lineNumber + 1, 'B', lineNumber + 4, 'number',
                    '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')
        formatRange('A', startBound, 'A', endBound, 'number', 'm/d/yy')
        formatRange('C', startBound, 'D', endBound, 'number', '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)')

        # Update line
        anchor = lineNumber + 4
        lineNumber = endBound + 2
    try:
        wb.save(settings['Out'])
    except:
        return 'Output file is currently open.'
    return ''

def createRow(rowCounter, start, end):
    startTuple = [int(x) for x in start.split('-')]
    endTuple = [int(x) for x in end.split('-')]
    return [
        sg.pin(
            sg.Col([[
                sg.Text("Start date:", tooltip='Start of a period. Should be YYYY-MM-DD'),
                sg.In(enable_events=True, key=('Start', rowCounter), default_text=start, size=(20, 1)),
                sg.CalendarButton("Select Date", format='%Y-%m-%d', size=(10, 1),
                                  default_date_m_d_y=(startTuple[1], startTuple[2], startTuple[0]),
                                  key=('CStart', rowCounter)),
                sg.Text("End date:", tooltip='End of a period. Should be YYYY-MM-DD'),
                sg.In(enable_events=True, key=('End', rowCounter), default_text=end, size=(20, 1)),
                sg.CalendarButton("Select Date", format='%Y-%m-%d', size=(10, 1),
                                  default_date_m_d_y=(endTuple[1], endTuple[2], endTuple[0]),
                                  key=('CEnd', rowCounter)),
                sg.Button('X', key=('Delete', rowCounter)),
            ]], key=('Row', rowCounter)),
        )
    ]

def main():
    # Loads inputs
    exists = os.path.isfile(settingsPath)
    settings = sg.UserSettings(filename=settingsPath, path='.')
    if not exists:
        defaultSettings()

    # Formats input window
    rows = []
    indices = []
    rowCounter = 0
    for period in settings['Periods']:
        rows.append(createRow(rowCounter, period[0], period[1]))
        indices.append(rowCounter)
        rowCounter += 1

    layout = [
        [
            sg.Column([
                [sg.Text("Input path:", pad=(0, 7),
                            tooltip='Path to a file with transaction data.\nFile should be a .csv and should have no more than one header.')],
                [sg.Text("Output path:", pad=(0, 7),
                            tooltip='Should be a .xlsx file. Creates one if it doesn\'t exist.')],
                [sg.Text("Excluded keywords:", pad=(0, 7),
                            tooltip='Any transaction description that contains these keywords will be kept\nas separate line items. Keywords are separated by commas. Ignores case.')],
                [sg.Text("Input date format:", pad=(0, 7),
                            tooltip='Date format in the input file.')],
            ]),
            sg.Column([
                [
                    sg.In(enable_events=True, key="In", default_text=settings['In']),
                    sg.FileBrowse(file_types=(("Comma Delimited", "*.csv"),), initial_folder='./'),
                ], [
                    sg.In(enable_events=True, key="Out", default_text=settings['Out']),
                    sg.FileSaveAs(file_types=(("Excel File", "*.xlsx"),), initial_folder='./'),
                ], [
                    sg.In(enable_events=True, key="Exclusions", default_text=settings['Exclusions']),
                ], [
                    sg.Combo(['%Y %m %d', '%d %m %Y', '%m %d %Y'],
                                enable_events=True, key="DateFormat", default_value=settings['DateFormat'], size=(10, 1)),
                ]
            ])
        ], [
            sg.Text("Post Date Column:"),
            sg.Combo(uppercase, enable_events=True, default_value=settings['Columns']['Post'], key='Post'),
            sg.Text("Description Column:"),
            sg.Combo(uppercase, enable_events=True, default_value=settings['Columns']['Description'], key='Description'),
            sg.Text("Debit Column:"),
            sg.Combo(uppercase, enable_events=True, default_value=settings['Columns']['Debit'], key='Debit'),
            sg.Text("Credit Column:"),
            sg.Combo(uppercase, enable_events=True, default_value=settings['Columns']['Credit'], key='Credit'),
        ], [
            sg.Text("Has Header:", tooltip='Check to ignore the first line.'),
            sg.Checkbox('', enable_events=True, default=settings['HasHeader'], key='HasHeader'),
            sg.Text("Credit card:", tooltip='Check if debits are positive additions to the account balance.\nThis is usually the case with credit cards.'),
            sg.Checkbox('', enable_events=True, default=settings['PositiveDebits'], key='PositiveDebits'),
            sg.Text("Sort by Description:", tooltip='Sort line items alphabetically by description.\nIf not, sort by date.'),
            sg.Checkbox('', enable_events=True, default=settings['SortDesc'], key='SortDesc'),
        ], [
            sg.Column(rows, key='RowPanel', scrollable = True, vertical_scroll_only = True, justification='center', size=(1000, 400))
        ], [
            sg.Button("Run"),
            sg.Button("Exit"),
            sg.Button("Add Date Range"),
            sg.Text('', key='IO'),
        ]
    ]
    window = sg.Window(title="Transaction Condenser",
                        layout=layout,
                        margins=(25, 25),
                        default_element_size=(80, 1),
                        font='Helvetica 15')

    def manualSave():
        settings.save(filename=settingsPath, path='./')

    # Create an event loop
    while True:
        event, values = window.read()

        if event in ['In', 'Out', 'DateFormat', 'HasHeader', 'Exclusions', 'PositiveDebits', 'SortDesc']:
            if isinstance(values[event], str):
                settings[event] = values[event].strip()
            else:
                settings[event] = values[event]
        elif event in ['Post', 'Description', 'Debit', 'Credit']:
            settings['Columns'][event] = values[event].strip()
            manualSave()
        elif event == 'Run':
            window['IO'].update(f'Running...')
            window.refresh()
            msg = parse()
            if msg:
                window['IO'].update(f'ERROR: {msg}')
            else:
                window['IO'].update('Successfully created the output file.')
        elif event == "Exit" or event == sg.WIN_CLOSED:
            break
        elif event == 'Add Date Range':
            try:
                start = datetime.strptime(settings['Periods'][-1][0], '%Y-%m-%d') + relativedelta(months=1)
                b = datetime.strptime(settings['Periods'][-1][1], '%Y-%m-%d')
                end = b + relativedelta(months=1)
                if b.month != (b + timedelta(days=1)).month: # Check if last day of month
                    end += relativedelta(day=31)
                start = datetime.strftime(start, '%Y-%m-%d')
                end = datetime.strftime(end, '%Y-%m-%d')
            except Exception as e:
                print(e)
                start = defaultStart
                end = defaultEnd
            window.extend_layout(window['RowPanel'], [createRow(rowCounter, start, end)])
            settings['Periods'].append([start, end])
            manualSave()
            indices.append(rowCounter)
            rowCounter += 1
            window.refresh()
            window['RowPanel'].contents_changed()
            window['IO'].update('Successfully added a date range.')
        elif event[0] == 'Delete':
            window[('Row', event[1])].update(visible=False)
            del settings['Periods'][indices.index(event[1])]
            manualSave()
            indices.remove(event[1])
            window.refresh()
            window['RowPanel'].contents_changed()
            window['IO'].update('Successfully deleted a date range.')
        elif event[0] in ['Start', 'End']:
            i = indices.index(event[1])
            newDate = values[(event[0], event[1])].strip()
            try:
                newTuple = [int(x) for x in newDate.split('-')]
                window[(f'C{event[0]}', event[1])].calendar_default_date_M_D_Y = (newTuple[1], newTuple[2], newTuple[0])
            except:
                pass
            settings['Periods'][i][event[0] != 'Start'] = newDate
            manualSave()
    window.close()

if __name__ == "__main__":
    main()
